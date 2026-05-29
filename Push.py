import json
import os
import socket
from datetime import datetime
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

FALLBACK_GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbyCK7M57DUQS6-KRJ5ZXQ-xjyj185IX_VZRaAtCVg6pakAy9gd4wvBnibmPPXWDq_sd/exec"
FRONTEND_ENV_FILE = os.path.join(os.path.dirname(__file__), "frontend", ".env")

KEY_FIELD = "APPLICATION_NUMBER_ID"
LOCAL_DATA_FILE = r"C:\Users\pengkheang.kang\Downloads\LoanRequest25052026084859.xlsx"
REQUEST_TIMEOUT_SECONDS = 360

# These columns are owned by the dashboard UI, so this local Excel push does not overwrite them.
LOCAL_ONLY_COLUMNS = [

    "FOLLOW_UP_REMARK",
    "REMARK_UPDATED_BY",
    "REMARK_UPDATED_AT",
]


def normalize(value):
    if value is None:
        return ""

    return str(value).strip()


def get_application_id(row):
    return normalize(row.get(KEY_FIELD))


def load_google_script_url():
    if not os.path.exists(FRONTEND_ENV_FILE):
        return FALLBACK_GOOGLE_SCRIPT_URL

    with open(FRONTEND_ENV_FILE, encoding="utf-8") as env_file:
        for line in env_file:
            key, separator, value = line.strip().partition("=")

            if separator and key == "VITE_SHEET_API_URL" and value:
                return value

    return FALLBACK_GOOGLE_SCRIPT_URL


def is_valid_application_row(row):
    application_id = get_application_id(row)
    return bool(application_id) and application_id != KEY_FIELD


def load_excel_rows():
    if not os.path.exists(LOCAL_DATA_FILE):
        raise FileNotFoundError(f"Local Excel file not found: {LOCAL_DATA_FILE}")

    workbook = load_workbook(LOCAL_DATA_FILE, data_only=True)
    sheet = workbook.active

    header_positions = []

    for column_index, cell in enumerate(sheet[1], start=1):
        header = normalize(cell.value)

        if header:
            header_positions.append((column_index, header))

    headers = [header for _, header in header_positions]

    if KEY_FIELD not in headers:
        raise RuntimeError(f"{KEY_FIELD} column not found in local Excel file.")

    rows = []

    for row_number in range(2, sheet.max_row + 1):
        row = {}

        for column_index, header in header_positions:
            value = sheet.cell(row=row_number, column=column_index).value
            row[header] = "" if value is None else value

        if is_valid_application_row(row):
            rows.append(row)

    return headers, rows


def build_columns_to_update(headers):
    columns = []

    for header in headers:
        if not header:
            continue

        if header in LOCAL_ONLY_COLUMNS:
            continue

        if header not in columns:
            columns.append(header)

    if KEY_FIELD not in columns:
        columns.insert(0, KEY_FIELD)

    return columns


def post_rows_to_google_sheet(rows, columns_to_update):
    google_script_url = load_google_script_url()

    payload = {
        "action": "syncRows",
        "keyField": KEY_FIELD,
        "columnsToUpdate": columns_to_update,
        "rows": rows,
    }

    body = json.dumps(payload, default=str).encode("utf-8")

    request = Request(
        google_script_url,
        data=body,
        method="POST",
        headers={
            "Content-Type": "text/plain;charset=utf-8",
        },
    )

    try:
        with urlopen(request, timeout=REQUEST_TIMEOUT_SECONDS) as response:
            raw = response.read().decode("utf-8")
    except HTTPError as exc:
        raise RuntimeError(f"Google Script HTTP error: {exc.code}") from exc
    except URLError as exc:
        raise RuntimeError(f"Cannot connect to Google Script: {exc.reason}") from exc
    except TimeoutError as exc:
        raise RuntimeError(
            f"Google Script did not respond within {REQUEST_TIMEOUT_SECONDS} seconds."
        ) from exc
    except socket.timeout as exc:
        raise RuntimeError(
            f"Google Script did not respond within {REQUEST_TIMEOUT_SECONDS} seconds."
        ) from exc

    result = json.loads(raw)

    if result.get("success") is False:
        raise RuntimeError(result.get("message", "Google Script returned an error"))

    return result


def main():
    print("Reading local Excel file...")
    headers, rows = load_excel_rows()
    columns_to_update = build_columns_to_update(headers)

    print(f"Local rows found: {len(rows)}")
    print(f"Columns to update in Google Sheet: {len(columns_to_update)}")
    print("Uploading local rows to Google Sheet...")

    result = post_rows_to_google_sheet(rows, columns_to_update)
    rows_to_modify = result.get("updated", 0) + result.get("appended", 0)

    print("\nGoogle Sheet updated.")
    print(f"LOS_Data rows that need to be modified: {rows_to_modify}")
    print(f"Updated existing rows: {result.get('updated', 0)}")
    print(f"Appended new rows: {result.get('appended', 0)}")
    print(f"Skipped invalid rows: {result.get('skipped', 0)}")

    if "skippedDrawdown" not in result or "removedDrawdown" not in result:
        print("WARNING: Deployed Apps Script did not return drawdown counters.")
        print("Redeploy Code.gs as a new Web App version before running Push.py again.")

    print(f"Skipped drawdown rows: {result.get('skippedDrawdown', 0)}")
    print(f"Removed drawdown rows from LOS_Data: {result.get('removedDrawdown', 0)}")

    if "drawdownIdCount" in result:
        print(f"DD application numbers found: {result.get('drawdownIdCount', 0)}")
    if "losRowsChecked" in result:
        print(f"LOS_Data rows checked for drawdown: {result.get('losRowsChecked', 0)}")
    if result.get("matchedDrawdownExamples"):
        examples = ", ".join(result["matchedDrawdownExamples"])
        print(f"Matched drawdown examples: {examples}")

    print(f"Updated at: {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
